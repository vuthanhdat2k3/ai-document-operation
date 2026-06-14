from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.cell.read_only import ReadOnlyCell
from openpyxl.worksheet.worksheet import Worksheet

from app.processing.parsers.base import (
    BaseParser,
    ImageData,
    PageResult,
    ParseResult,
    TableData,
)

logger = logging.getLogger(__name__)


class XLSXParser(BaseParser):
    """Parse XLSX spreadsheets using openpyxl.

    Capabilities:
    - Sheet-by-sheet extraction
    - Cell value extraction with type preservation
    - Merged-cell handling (top-left value is used)
    - Formula extraction as text (e.g. ``=SUM(A1:A10)``)
    """

    def parse(self, file_path: Path) -> ParseResult:
        if not file_path.exists():
            raise FileNotFoundError(f"XLSX not found: {file_path}")

        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=False)
        except Exception as exc:
            raise ValueError(f"Cannot open XLSX: {exc}") from exc

        pages: list[PageResult] = []
        total_cells = 0
        non_empty_cells = 0

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                page_result = self._parse_sheet(ws, sheet_name, len(pages) + 1)
                pages.append(page_result)

                for table in page_result.tables:
                    total_cells += len(table.headers)
                    for row in table.rows:
                        total_cells += len(row)
                        non_empty_cells += sum(1 for c in row if c.strip())
                    non_empty_cells += sum(1 for h in table.headers if h.strip())

            except Exception:
                logger.warning("Failed to parse sheet '%s'", sheet_name, exc_info=True)
                pages.append(
                    PageResult(
                        page_number=len(pages) + 1,
                        text=f"[Error parsing sheet: {sheet_name}]",
                        confidence=0.0,
                    )
                )

        metadata = self._extract_metadata(wb)
        wb.close()

        total_pages = len(pages)
        avg_confidence = (
            sum(p.confidence for p in pages) / total_pages if total_pages else 0.0
        )

        return ParseResult(
            pages=pages,
            metadata=metadata,
            quality_score=round(avg_confidence, 4),
        )

    def _parse_sheet(
        self, ws: Worksheet, sheet_name: str, page_number: int
    ) -> PageResult:
        rows_data: list[list[str]] = []
        max_col = 0

        merged_map = self._build_merged_map(ws)

        for row in ws.iter_rows():
            row_values: list[str] = []
            for cell in row:
                value = self._cell_to_string(cell, merged_map)
                row_values.append(value)
            # Trim trailing empty cells
            while row_values and not row_values[-1].strip():
                row_values.pop()
            if row_values:
                max_col = max(max_col, len(row_values))
            rows_data.append(row_values)

        # Remove trailing empty rows
        while rows_data and not any(c.strip() for c in rows_data[-1]):
            rows_data.pop()

        # Normalise column counts
        for row in rows_data:
            while len(row) < max_col:
                row.append("")

        if not rows_data:
            return PageResult(
                page_number=page_number,
                text=f"[Sheet: {sheet_name} — empty]",
                tables=[],
                confidence=1.0,
            )

        headers = rows_data[0]
        data_rows = rows_data[1:] if len(rows_data) > 1 else []
        table = TableData(headers=headers, rows=data_rows, page=page_number)

        non_empty = sum(
            1
            for row in rows_data
            for cell in row
            if cell.strip()
        )
        total = len(rows_data) * max_col if max_col else 1
        confidence = min(1.0, non_empty / max(total, 1) * 2)

        text_summary = self._build_text_summary(sheet_name, rows_data)

        return PageResult(
            page_number=page_number,
            text=text_summary,
            tables=[table],
            confidence=round(confidence, 4),
        )

    @staticmethod
    def _build_merged_map(ws: Worksheet) -> dict[tuple[int, int], str]:
        """Map every cell in a merged range to the top-left cell's value."""
        merged: dict[tuple[int, int], str] = {}
        for rng in ws.merged_cells.ranges:
            try:
                top_left = ws.cell(row=rng.min_row, column=rng.min_col)
                value = str(top_left.value) if top_left.value is not None else ""
            except Exception:
                value = ""
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    merged[(row, col)] = value
        return merged

    @staticmethod
    def _cell_to_string(
        cell: Cell | ReadOnlyCell | MergedCell,
        merged_map: dict[tuple[int, int], str],
    ) -> str:
        key = (cell.row, cell.column)

        if isinstance(cell, MergedCell):
            return merged_map.get(key, "")

        value = cell.value
        if value is None:
            return merged_map.get(key, "")

        # Show formulas as text
        if isinstance(value, str) and value.startswith("="):
            return value

        return str(value)

    @staticmethod
    def _build_text_summary(sheet_name: str, rows: list[list[str]]) -> str:
        lines = [f"[Sheet: {sheet_name}]"]
        max_preview = min(len(rows), 50)
        for row in rows[:max_preview]:
            line = " | ".join(row)
            if line.strip():
                lines.append(line)
        if len(rows) > max_preview:
            lines.append(f"... ({len(rows) - max_preview} more rows)")
        return "\n".join(lines)

    @staticmethod
    def _extract_metadata(wb) -> dict:
        try:
            props = wb.properties
        except Exception:
            return {}

        return {
            "title": props.title or "",
            "creator": props.creator or "",
            "subject": props.subject or "",
            "description": props.description or "",
            "created": str(props.created) if props.created else "",
            "modified": str(props.modified) if props.modified else "",
            "sheet_names": list(wb.sheetnames),
            "sheet_count": len(wb.sheetnames),
        }
